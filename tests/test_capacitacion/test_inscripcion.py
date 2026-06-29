import io
from datetime import date, time, timedelta

import pytest
from django.core import mail
from django.urls import reverse
from openpyxl import Workbook

from apps.capacitacion.models import Campana, Inscripcion, Sede, Sesion
from apps.capacitacion.services import (
    generar_plantilla_excel,
    generar_qr_png,
    importar_excel_completo,
    inscripcion_url_por_sede,
)
from apps.participantes.models import Participante


@pytest.fixture
def campana(db):
    return Campana.objects.create(nombre="Campaña test", activa=True)


@pytest.fixture
def sede(db, campana):
    return Sede.objects.create(
        campana=campana,
        nombre="UCV Test",
        slug="ucv-test",
        activa=True,
    )


@pytest.fixture
def sesion(db, sede):
    return Sesion.objects.create(
        sede=sede,
        fecha=date.today() + timedelta(days=7),
        hora=time(9, 0),
        cupo=50,
    )


@pytest.mark.django_db
class TestInscripcionPublica:
    def test_inicio_inscripcion_ok(self, client, sede, sesion):
        r = client.get(reverse("capacitacion_public:paso1"))
        assert r.status_code == 200

    def test_inscripcion_por_sede_qr(self, client, sede, sesion):
        r = client.get(reverse("capacitacion_public:por_sede", kwargs={"slug": sede.slug}))
        assert r.status_code == 200
        assert sede.nombre in r.content.decode()

    def test_flujo_inscripcion_completo(self, client, sede, sesion):
        url_paso2 = reverse("capacitacion_public:paso2", kwargs={"sesion_id": sesion.pk})
        r = client.post(
            url_paso2,
            {
                "cedula": "V12345678",
                "apellidos": "Pérez",
                "nombres": "Juan",
                "telefono": "04141234567",
                "correo": "juan@example.com",
                "profesion": Participante.Profesion.INGENIERO_CIVIL,
                "profesion_otro": "",
                "procedencia": "UCV",
            },
        )
        assert r.status_code == 302
        ins = Inscripcion.objects.get(participante__cedula="V12345678")
        assert ins.codigo.startswith("INS-")
        assert len(mail.outbox) == 1
        assert "juan@example.com" in mail.outbox[0].to

    def test_confirmacion_muestra_codigo(self, client, sede, sesion):
        p = Participante.objects.create(
            cedula="V87654321",
            apellidos="Gómez",
            nombres="Ana",
            telefono="04140001111",
            procedencia="CIV",
        )
        ins = Inscripcion.objects.create(sesion=sesion, participante=p)
        r = client.get(reverse("capacitacion_public:confirmacion", kwargs={"codigo": ins.codigo}))
        assert r.status_code == 200
        assert ins.codigo in r.content.decode()


@pytest.mark.django_db
class TestExcelImport:
    def test_plantilla_tiene_hojas(self):
        data = generar_plantilla_excel()
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data))
        assert "sedes" in wb.sheetnames
        assert "sesiones" in wb.sheetnames

    def test_importar_sedes_y_sesiones(self, campana):
        wb = Workbook()
        ws = wb.active
        ws.title = "sedes"
        ws.append(["nombre", "slug", "direccion", "municipio", "estado", "responsable"])
        ws.append(["Sede Nueva", "sede-nueva", "Dir", "Caracas", "DC", "Coord"])
        ws2 = wb.create_sheet("sesiones")
        ws2.append(["sede_slug", "fecha", "hora", "cupo"])
        ws2.append(["sede-nueva", "2026-08-01", "10:00", 60])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = importar_excel_completo(buf, campana=campana)
        assert result["sedes"]["creadas"] == 1
        assert result["sesiones"]["creadas"] == 1
        assert Sede.objects.filter(slug="sede-nueva").exists()


@pytest.mark.django_db
class TestQR:
    def test_genera_png(self, sede):
        url = inscripcion_url_por_sede(sede)
        png = generar_qr_png(url)
        assert png[:8] == b"\x89PNG\r\n\x1a\n"


@pytest.mark.django_db
class TestPanelCoordinador:
    def test_panel_requiere_login(self, client):
        r = client.get(reverse("capacitacion:panel_inicio"))
        assert r.status_code == 302

    def test_panel_staff_ok(self, client, admin_user):
        client.force_login(admin_user)
        r = client.get(reverse("capacitacion:panel_inicio"))
        assert r.status_code == 200

    def test_descarga_qr(self, client, admin_user, sede):
        client.force_login(admin_user)
        r = client.get(reverse("capacitacion:panel_sede_qr", kwargs={"sede_id": sede.pk}))
        assert r.status_code == 200
        assert r["Content-Type"] == "image/png"
