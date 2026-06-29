"""Utilidades compartidas del panel de capacitación."""
from __future__ import annotations

import io
from datetime import date, datetime, time
from typing import Any

from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string
from openpyxl import Workbook, load_workbook

from apps.capacitacion.models import Campana, Inscripcion, Sede, Sesion


def site_base_url(request=None) -> str:
    if request:
        return request.build_absolute_uri("/").rstrip("/")
    return getattr(settings, "SITE_URL", "http://127.0.0.1:8000").rstrip("/")


def inscripcion_url_por_sede(sede: Sede, request=None) -> str:
    return f"{site_base_url(request)}/inscripcion/sede/{sede.slug}/"


def generar_qr_png(url: str) -> bytes:
    import qrcode

    qr = qrcode.QRCode(version=1, box_size=8, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#00247D", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    return buffer.getvalue()


def enviar_correo_confirmacion(inscripcion: Inscripcion, request=None) -> bool:
    participante = inscripcion.participante
    if not participante.correo:
        return False

    context = {
        "inscripcion": inscripcion,
        "participante": participante,
        "sesion": inscripcion.sesion,
        "sede": inscripcion.sesion.sede,
        "comision": settings.CPEH_NOMBRE_COMISION,
        "evento_fecha": settings.CPEH_EVENTO_SISMO_FECHA,
    }
    subject = f"Inscripción confirmada — {inscripcion.codigo}"
    body = render_to_string("capacitacion/email/confirmacion_inscripcion.txt", context)
    html_body = render_to_string("capacitacion/email/confirmacion_inscripcion.html", context)

    send_mail(
        subject=subject,
        message=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[participante.correo],
        html_message=html_body,
        fail_silently=False,
    )
    return True


def _parse_fecha(valor: Any) -> date:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha no válida: {valor}")


def _parse_hora(valor: Any) -> time:
    if isinstance(valor, datetime):
        return valor.time().replace(second=0, microsecond=0)
    if isinstance(valor, time):
        return valor.replace(second=0, microsecond=0)
    texto = str(valor).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(texto, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Hora no válida: {valor}")


def _cell(row: tuple, idx: int) -> str:
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def importar_sedes_desde_excel(archivo, campana: Campana | None = None) -> dict[str, int]:
    wb = load_workbook(archivo, read_only=True, data_only=True)
    if "sedes" not in wb.sheetnames:
        raise ValueError('El archivo debe tener una hoja llamada "sedes".')
    ws = wb["sedes"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    creadas = actualizadas = omitidas = 0

    for row in rows:
        if not row or not _cell(row, 0):
            continue
        nombre = _cell(row, 0)
        slug = _cell(row, 1) or None
        direccion = _cell(row, 2)
        municipio = _cell(row, 3)
        estado = _cell(row, 4)
        responsable = _cell(row, 5)

        lookup = {"slug": slug} if slug else {"nombre": nombre}
        defaults = {
            "nombre": nombre,
            "direccion": direccion,
            "municipio": municipio,
            "estado": estado,
            "responsable_nombre": responsable,
            "activa": True,
            "campana": campana,
        }
        if slug:
            sede, created = Sede.objects.update_or_create(slug=slug, defaults=defaults)
        else:
            sede, created = Sede.objects.get_or_create(nombre=nombre, defaults=defaults)
            if not created:
                for k, v in defaults.items():
                    setattr(sede, k, v)
                sede.save()

        if created:
            creadas += 1
        else:
            actualizadas += 1

    return {"creadas": creadas, "actualizadas": actualizadas, "omitidas": omitidas}


def importar_sesiones_desde_excel(archivo) -> dict[str, int]:
    wb = load_workbook(archivo, read_only=True, data_only=True)
    if "sesiones" not in wb.sheetnames:
        raise ValueError('El archivo debe tener una hoja llamada "sesiones".')
    ws = wb["sesiones"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    creadas = omitidas = errores = 0
    detalle_errores: list[str] = []

    for i, row in enumerate(rows, start=2):
        if not row or not _cell(row, 0):
            continue
        sede_ref = _cell(row, 0)
        try:
            sede = Sede.objects.filter(slug=sede_ref).first() or Sede.objects.filter(nombre=sede_ref).first()
            if not sede:
                raise ValueError(f"Sede no encontrada: {sede_ref}")
            fecha = _parse_fecha(row[1])
            hora = _parse_hora(row[2])
            cupo = int(row[3]) if len(row) > 3 and row[3] not in (None, "") else 80
            _, created = Sesion.objects.get_or_create(
                sede=sede,
                fecha=fecha,
                hora=hora,
                defaults={"cupo": cupo},
            )
            if created:
                creadas += 1
            else:
                omitidas += 1
        except Exception as exc:
            errores += 1
            detalle_errores.append(f"Fila {i}: {exc}")

    result = {"creadas": creadas, "omitidas": omitidas, "errores": errores}
    if detalle_errores:
        result["detalle_errores"] = detalle_errores[:20]
    return result


def importar_excel_completo(archivo, campana: Campana | None = None) -> dict:
    resultado = {"sedes": None, "sesiones": None}
    wb = load_workbook(archivo, read_only=True, data_only=True)
    archivo.seek(0)
    if "sedes" in wb.sheetnames:
        resultado["sedes"] = importar_sedes_desde_excel(archivo, campana=campana)
        archivo.seek(0)
    if "sesiones" in wb.sheetnames:
        resultado["sesiones"] = importar_sesiones_desde_excel(archivo)
    if not resultado["sedes"] and not resultado["sesiones"]:
        raise ValueError('El Excel debe incluir al menos una hoja "sedes" o "sesiones".')
    return resultado


def generar_plantilla_excel() -> bytes:
    wb = Workbook()
    ws_sedes = wb.active
    ws_sedes.title = "sedes"
    ws_sedes.append(["nombre", "slug", "direccion", "municipio", "estado", "responsable"])
    ws_sedes.append([
        "Universidad Central de Venezuela (UCV)",
        "ucv",
        "Los Chaguaramos",
        "Caracas",
        "Distrito Capital",
        "Ing. Coordinador",
    ])

    ws_sesiones = wb.create_sheet("sesiones")
    ws_sesiones.append(["sede_slug", "fecha", "hora", "cupo"])
    ws_sesiones.append(["ucv", "2026-07-01", "09:00", 100])
    ws_sesiones.append(["ucv", "2026-07-02", "09:00", 100])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def exportar_inscritos_sesion(sesion: Sesion) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "inscritos"
    ws.append([
        "codigo",
        "cedula",
        "apellidos",
        "nombres",
        "telefono",
        "correo",
        "profesion",
        "procedencia",
        "sede",
        "fecha",
        "hora",
        "asistio",
        "certificado",
    ])
    qs = (
        Inscripcion.objects.filter(sesion=sesion)
        .select_related("participante", "sesion__sede")
        .order_by("participante__apellidos")
    )
    for ins in qs:
        p = ins.participante
        ws.append([
            ins.codigo,
            p.cedula,
            p.apellidos,
            p.nombres,
            p.telefono,
            p.correo,
            p.profesion_display(),
            p.procedencia,
            sesion.sede.nombre,
            sesion.fecha.isoformat(),
            sesion.hora.strftime("%H:%M"),
            "Sí" if ins.asistio else "No",
            "Sí" if ins.certificado_emitido else "No",
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
